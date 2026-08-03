import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

rows = []
with open('loto_ci_final.csv', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for r in reader:
        rows.append(r)

wb = Workbook()
ws = wb.active
ws.title = "Historique LONACI"

headers = ["Date", "Jour", "Jeu", "Numéros gagnants", "Numéros machine"]
ws.append(headers)

header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
for col_idx, h in enumerate(headers, start=1):
    c = ws.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")

data_font = Font(name="Arial", size=11)

for r in rows:
    day, month, year = r['Date'].split('/')
    from datetime import date
    d = date(int(year), int(month), int(day))
    ws.append([d, r['Jour'], r['Jeu'], r['Numeros'], r['Machine'] if r['Machine'] else None])

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
    for cell in row:
        cell.font = data_font
        if cell.column == 1:
            cell.number_format = "DD/MM/YYYY"
            cell.alignment = Alignment(horizontal="center")
        elif cell.column in (4,5):
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.alignment = Alignment(horizontal="left")

widths = {1: 13, 2: 12, 3: 20, 4: 20, 5: 20}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:E{ws.max_row}"

wb.save("Historique Loto Ivoirien (LONACI).xlsx")
print("Saved. Rows written (excl header):", ws.max_row - 1)
